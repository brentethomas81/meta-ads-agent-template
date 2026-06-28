#!/usr/bin/env python3
"""Build the Your Brand AD CAMPAIGN tracker from adtracker.json.
Usage: python3 build_ad_tracker_xlsx.py <adtracker.json> <out.xlsx>
adtracker.json: {"meta":{spend,impressions,reach,frequency,clicks,link_clicks,ctr,cpc,cpm,lpv,start},
                 "ad_spend":float (account lifetime), "ad_subs":[{email,plan,amount,interval,status,start,canceled,collected,tier}],
                 "generated":str}
Two attribution tiers per customer: CONFIRMED (hard click id / the /your-landing landing page / hand-confirmed)
and LIKELY (a first-time visitor who bought the ad's plan minutes after landing — click id wiped by the
in-app browser). ROI is shown both ways. Focused on the AD CAMPAIGN only. Pure openpyxl (no recalc)."""
import json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference

inp = sys.argv[1] if len(sys.argv) > 1 else "adtracker.json"
out = sys.argv[2] if len(sys.argv) > 2 else "AdCampaignTracker.xlsx"
d = json.load(open(inp)); m = d.get("meta", {}); gen = d.get("generated", "")
spend = float(m.get("spend") or 0)               # this campaign's spend (the right ROI denominator)
subs = [r for r in d.get("ad_subs", []) if r.get("status") not in ("—", None) and r.get("plan") not in ("(no Stripe customer)", "err")]

NAVY="1E293B"; PURPLE="7C3AED"; GREEN="16A34A"; RED="DC2626"; GREY="64748B"
LIGHT="F1F5F9"; LAV="F3F1FF"; WHITE="FFFFFF"; AMBER="D97706"; BLUE="1877F2"; FONT="Arial"
def Fn(**k): return Font(name=FONT, **k)
thin=Side(style="thin", color="E2E8F0"); box=Border(left=thin,right=thin,top=thin,bottom=thin)
wb=Workbook()

# ---- Ad Customers tab ----
ws=wb.create_sheet("Ad Customers")
ws.append(["Customer","Plan","Started","Status","Canceling","Paid so far","Tier"])
smap={"active":"Active","canceled":"Canceled","past_due":"Past due","trialing":"Trialing"}
tmap={"confirmed":"Confirmed","likely":"Likely"}
for r in subs:
    ws.append([r["email"], r.get("plan",""), r.get("start",""), smap.get(r.get("status"), str(r.get("status")).title()),
               r.get("canceled",""), r.get("collected",0), tmap.get(r.get("tier"), str(r.get("tier","")).title())])
last=len(subs)+1
for c in range(1,8):
    cell=ws.cell(1,c); cell.font=Fn(bold=True,color=WHITE,size=11); cell.fill=PatternFill("solid",fgColor=NAVY)
    cell.alignment=Alignment(horizontal="center",vertical="center"); cell.border=box
ws.row_dimensions[1].height=22
for i,w in zip("ABCDEFG",[34,16,12,12,12,12,12]): ws.column_dimensions[i].width=w
for i in range(2,last+1):
    st=ws[f"D{i}"].value
    ws[f"D{i}"].font=Fn(color=GREEN if st=="Active" else RED, bold=True, size=10)
    ws[f"F{i}"].number_format='$#,##0.00'
    tv=ws[f"G{i}"].value
    ws[f"G{i}"].font=Fn(color=GREEN if tv=="Confirmed" else AMBER, bold=True, size=10)
    for c in range(1,8):
        cell=ws.cell(i,c); cell.border=box
        cell.alignment=Alignment(horizontal="left" if c==1 else "center", vertical="center")
        if c not in (4,7): cell.font=Fn(size=10)
ws.freeze_panes="A2"

# ---- Campaign dashboard ----
db=wb["Sheet"]; db.title="Ad Campaign"; db.sheet_view.showGridLines=False
db["A1"]="Your Brand — Ad Campaign Tracker"; db["A1"].font=Fn(bold=True,size=20,color=NAVY)
db["A2"]=f"@your-ig → /your-landing · running since {m.get('start','')} · updated {gen}"; db["A2"].font=Fn(size=10,color=GREY)
db.merge_cells("A1:H1"); db.merge_cells("A2:H2")
def card(anchor,label,val,fmt,color,fill):
    col=anchor[0]; row=int(anchor[1:]); nx=chr(ord(col)+1)
    db[f"{col}{row}"].value=label; db[f"{col}{row}"].font=Fn(size=9,bold=True,color=GREY)
    v=db[f"{col}{row+1}"]; v.value=val; v.font=Fn(size=18,bold=True,color=color); v.number_format=fmt
    db.merge_cells(f"{col}{row}:{nx}{row}"); db.merge_cells(f"{col}{row+1}:{nx}{row+1}")
    for cc in (col,nx):
        for rr in (row,row+1):
            db[f"{cc}{rr}"].fill=PatternFill("solid",fgColor=fill); db[f"{cc}{rr}"].border=box
    db[f"{col}{row}"].alignment=Alignment(horizontal="left",vertical="center")
    db[f"{col}{row+1}"].alignment=Alignment(horizontal="left",vertical="center")

# spend lives in J2 as the editable ROI denominator
db["I2"]="Ad spend (edit) →"; db["I2"].font=Fn(size=9,color=GREY); db["I2"].alignment=Alignment(horizontal="right")
db["J2"]=round(spend,2); db["J2"].font=Fn(size=11,bold=True,color="0000FF"); db["J2"].number_format='$#,##0.00'

db["A4"]="AD DELIVERY  (what Meta did)"; db["A4"].font=Fn(bold=True,size=12,color=NAVY)
card("A5","SPENT",round(spend,2),"$#,##0.00",AMBER,LIGHT)
card("C5","PEOPLE REACHED",int(m.get("reach") or 0),"#,##0",NAVY,LIGHT)
card("E5","LINK CLICKS",int(m.get("link_clicks") or 0),"#,##0",NAVY,LIGHT)
card("G5","CTR (% who click)",round((m.get('ctr') or 0)/100,4),"0.0%",GREEN,LIGHT)
card("A7","COST PER CLICK",round(m.get('cpc') or 0,2),"$#,##0.00",NAVY,LIGHT)
card("C7","LANDING-PAGE VIEWS",int(m.get("lpv") or 0),"#,##0",NAVY,LIGHT)

S="'Ad Customers'"; F=f"{S}!F2:F{last}"; D=f"{S}!D2:D{last}"; T=f"{S}!G2:G{last}"
CONF_REV=f'SUMIFS({F},{T},"Confirmed")'
ALL_REV=f"SUM({F})"
db["A10"]="RESULTS & ROI  (what the ad earned — two confidence tiers)"; db["A10"].font=Fn(bold=True,size=12,color=PURPLE)
card("A11","AD CUSTOMERS (total)",f"=COUNTA({S}!A2:A{last})","#,##0",PURPLE,LAV)
card("C11","CONFIRMED",f'=COUNTIF({T},"Confirmed")',"#,##0",GREEN,LAV)
card("E11","LIKELY",f'=COUNTIF({T},"Likely")',"#,##0",AMBER,LAV)
card("G11","STILL ACTIVE",f'=COUNTIF({D},"Active")',"#,##0",GREEN,LAV)
card("A13","COLLECTED (all believed)",f"={ALL_REV}","$#,##0.00",GREEN,LAV)
card("C13","ROI — confirmed only",f"=IF(J2=0,0,{CONF_REV}/J2)","0%",GREEN,LAV)
card("E13","ROI — incl. likely",f"=IF(J2=0,0,{ALL_REV}/J2)","0%",GREEN,LAV)
card("G13","$ BACK PER $1 (all)",f"=IF(J2=0,0,{ALL_REV}/J2)","$#,##0.00",GREEN,LAV)
db["A15"]=("Confirmed = hard proof (Meta click id / arrived via the /your-landing page / hand-verified).  "
           "Likely = first-time visitor who bought the your plan minutes after landing (the in-app browser strips the click id, so real ad sales often arrive bare).")
db["A15"].font=Fn(size=9,italic=True,color=GREY); db.merge_cells("A15:H15"); db.row_dimensions[15].height=26
db["A16"]="Note: collected = payments so far. As these subs renew, ROI climbs."; db["A16"].font=Fn(size=9,italic=True,color=GREY); db.merge_cells("A16:H16")

# chart: spend vs collected
db["N20"]="Ad spend"; db["O20"]="=J2"; db["N21"]="Collected"; db["O21"]=f"={ALL_REV}"
bar=BarChart(); bar.title="Ad spend vs collected"; bar.type="col"; bar.legend=None; bar.height=7; bar.width=11
bar.add_data(Reference(db,min_col=15,min_row=20,max_row=21)); bar.set_categories(Reference(db,min_col=14,min_row=20,max_row=21))
db.add_chart(bar,"A18")
for col,w in zip("ABCDEFGH",[20,16,18,16,20,14,22,14]): db.column_dimensions[col].width=w
wb.save(out)
print("built ad tracker:", len(subs), "ad customers ->", out)
