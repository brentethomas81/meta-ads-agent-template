#!/usr/bin/env python3
"""Build the Your Brand 'Subscriptions & Ad ROI' workbook from a subs.json
dump. Usage: python3 build_tracker_xlsx.py <input subs.json> <output .xlsx>
subs.json shape: {"rows":[{email,plan,amount,interval,status,start,canceled,collected,is_ad}], "ad_spend":float, "generated":str}
Run scripts/recalc.py on the output afterward so the dashboard formulas cache values."""
import json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.utils import get_column_letter

inp = sys.argv[1] if len(sys.argv) > 1 else "subs.json"
out = sys.argv[2] if len(sys.argv) > 2 else "YourBrand_Subscriptions.xlsx"
d = json.load(open(inp)); rows = d["rows"]; spend = d["ad_spend"]; gen = d["generated"]

NAVY="1E293B"; PURPLE="7C3AED"; GREEN="16A34A"; RED="DC2626"; GREY="64748B"
LIGHT="F1F5F9"; LAV="F3F1FF"; WHITE="FFFFFF"; AMBER="D97706"; FONT="Arial"
def Fn(**k): return Font(name=FONT, **k)
thin=Side(style="thin", color="E2E8F0"); box=Border(left=thin,right=thin,top=thin,bottom=thin)
wb=Workbook()
ws=wb.create_sheet("Subscribers")
ws.append(["Email","Plan","Status","Started","Canceled","Total Paid","From Ad?","Amount","Interval","MRR"])
smap={"active":"Active","canceled":"Canceled","past_due":"Past due","trialing":"Trialing","unpaid":"Unpaid","incomplete":"Incomplete","paused":"Paused"}
for r in rows:
    ws.append([r["email"] or "(no email)", r["plan"], smap.get(r["status"], r["status"].title()),
               r["start"], r["canceled"], r["collected"], "Yes" if r["is_ad"] else "No",
               r["amount"], r["interval"], None])
n=len(rows); last=n+1
for i in range(2,last+1):
    ws[f"J{i}"]=f'=IF(I{i}="year",H{i}/12,IF(I{i}="month",H{i},0))'
    for c in ("F","H","J"): ws[f"{c}{i}"].number_format='$#,##0.00'
for c in range(1,11):
    cell=ws.cell(1,c); cell.font=Fn(bold=True,color=WHITE,size=11); cell.fill=PatternFill("solid",fgColor=NAVY)
    cell.alignment=Alignment(horizontal="center",vertical="center"); cell.border=box
ws.row_dimensions[1].height=22
for i,w in enumerate([34,14,12,12,12,13,10,10,10,11],1): ws.column_dimensions[get_column_letter(i)].width=w
for c in "HIJ": ws.column_dimensions[c].hidden=True
ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:G{last}"
for i in range(2,last+1):
    if i%2==0:
        for c in range(1,8): ws.cell(i,c).fill=PatternFill("solid",fgColor="FAFAFC")
    st=ws[f"C{i}"].value
    ws[f"C{i}"].font=Fn(color=GREEN if st=="Active" else (RED if st=="Canceled" else AMBER),bold=True,size=10)
    if ws[f"G{i}"].value=="Yes":
        ws[f"G{i}"].font=Fn(bold=True,color=PURPLE); ws[f"G{i}"].fill=PatternFill("solid",fgColor=LAV)
    for c in range(1,8):
        cell=ws.cell(i,c); cell.border=box
        if c in (1,2,4,5): cell.font=Fn(size=10)
        cell.alignment=Alignment(horizontal="left" if c==1 else "center",vertical="center")
db=wb["Sheet"]; db.title="Dashboard"; db.sheet_view.showGridLines=False
S="Subscribers"
db["A1"]="Your Brand — Subscriptions & Ad ROI"; db["A1"].font=Fn(bold=True,size=20,color=NAVY)
db["A2"]=f"Live from Stripe + Meta · generated {gen} · {n:,} subscribers on file"; db["A2"].font=Fn(size=10,color=GREY)
db.merge_cells("A1:H1"); db.merge_cells("A2:H2")
db["I2"]="Ad spend (edit) →"; db["I2"].font=Fn(size=9,color=GREY); db["I2"].alignment=Alignment(horizontal="right")
db["J2"]=spend; db["J2"].font=Fn(size=11,bold=True,color="0000FF"); db["J2"].number_format='$#,##0.00'
def card(anchor,label,val,fmt,color,fill):
    col=anchor[0]; row=int(anchor[1:]); nx=chr(ord(col)+1)
    db[f"{col}{row}"].value=label; db[f"{col}{row}"].font=Fn(size=9,bold=True,color=GREY)
    v=db[f"{col}{row+1}"]; v.value=val; v.font=Fn(size=18,bold=True,color=color); v.number_format=fmt
    db.merge_cells(f"{col}{row}:{nx}{row}"); db.merge_cells(f"{col}{row+1}:{nx}{row+1}")
    for cc in (col,nx):
        for rr in (row,row+1):
            db[f"{cc}{rr}"].fill=PatternFill("solid",fgColor=fill); db[f"{cc}{rr}"].border=box
    db[f"{col}{row}"].alignment=Alignment(horizontal="left",vertical="center")
    db[f"{col}{row+1}"].alignment=Alignment(horizontal="left",vertical="center")
A=f"{S}!C2:C{last}"; FC=f"{S}!F2:F{last}"; J=f"{S}!J2:J{last}"; G=f"{S}!G2:G{last}"; TOT=f"COUNTA({S}!C2:C{last})"
db["A4"]="THE BUSINESS"; db["A4"].font=Fn(bold=True,size=12,color=NAVY)
card("A5","TOTAL SUBSCRIBERS",f"={TOT}","#,##0",NAVY,LIGHT)
card("C5","ACTIVE",f'=COUNTIF({A},"Active")',"#,##0",GREEN,LIGHT)
card("E5","CANCELED (CHURNED)",f'=COUNTIF({A},"Canceled")',"#,##0",RED,LIGHT)
card("G5","CHURN RATE",f'=COUNTIF({A},"Canceled")/{TOT}',"0.0%",AMBER,LIGHT)
card("A7","MONTHLY RECURRING REVENUE",f'=SUMIF({A},"Active",{J})',"$#,##0.00",NAVY,LIGHT)
card("C7","COLLECTED (LIFETIME)",f"=SUM({FC})","$#,##0.00",GREEN,LIGHT)
db["A10"]="ADS — RETURN ON INVESTMENT"; db["A10"].font=Fn(bold=True,size=12,color=PURPLE)
card("A11","AD SPEND","=J2","$#,##0.00",AMBER,LAV)
card("C11","AD SUBSCRIBERS",f'=COUNTIF({G},"Yes")',"#,##0",PURPLE,LAV)
card("E11","AD MRR",f'=SUMIFS({J},{G},"Yes",{A},"Active")',"$#,##0.00",PURPLE,LAV)
card("G11","AD REVENUE COLLECTED",f'=SUMIF({G},"Yes",{FC})',"$#,##0.00",GREEN,LAV)
card("A13","ROI (revenue ÷ spend)",f'=IF(J2=0,0,SUMIF({G},"Yes",{FC})/J2)',"0%",GREEN,LAV)
card("C13","$ BACK PER $1 SPENT",f'=IF(J2=0,0,SUMIF({G},"Yes",{FC})/J2)',"$#,##0.00",GREEN,LAV)
db["N30"]="Active"; db["O30"]=f'=COUNTIF({A},"Active")'
db["N31"]="Canceled"; db["O31"]=f'=COUNTIF({A},"Canceled")'
db["N32"]="Other"; db["O32"]=f'={TOT}-O30-O31'
pie=PieChart(); pie.title="Subscriber status"; pie.height=7.5; pie.width=10.5
pie.add_data(Reference(db,min_col=15,min_row=30,max_row=32)); pie.set_categories(Reference(db,min_col=14,min_row=30,max_row=32))
db.add_chart(pie,"A16")
db["N34"]="Ad spend"; db["O34"]="=J2"; db["N35"]="Ad revenue"; db["O35"]=f'=SUMIF({G},"Yes",{FC})'
bar=BarChart(); bar.title="Ad spend vs revenue collected"; bar.type="col"; bar.legend=None; bar.height=7.5; bar.width=10.5
bar.add_data(Reference(db,min_col=15,min_row=34,max_row=35)); bar.set_categories(Reference(db,min_col=14,min_row=34,max_row=35))
db.add_chart(bar,"E16")
for col,w in zip("ABCDEFGH",[20,16,18,16,20,14,20,14]): db.column_dimensions[col].width=w
wb.save(out)
print("built", n, "rows ->", out)
